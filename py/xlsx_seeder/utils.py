import json
import os
import transform
from tqdm import tqdm
from openpyxl import load_workbook
from os.path import exists
import psycopg2

def get_connection_string(connection):
    f = open('settings.json', encoding='utf-8')
    data = json.load(f)
    conn_string = None
    if connection in data["connections"]:
        dbname = data["connections"][connection]["dbName"]
        user = data["connections"][connection]["user"]
        password = data["connections"][connection]["password"]
        host = data["connections"][connection]["host"]
        port = data["connections"][connection]["port"]
        conn_string = "dbname=%s user=%s password=%s host=%s port=%s" % (dbname, user, password,host, port)
    return conn_string

def get_connection(connection):
    conn_string = get_connection_string(connection)
    if conn_string is None:
        raise Exception("En el archivo de settings.json, no existe la conexión '" + connection + "'")
    #Envio a la base de datos    
    return psycopg2.connect(conn_string)

def load_excel(file, table):
    if not exists(file):
        raise Exception("No existe el archivo en la ruta; " + file)
    extension = get_extension(file)
    if extension not in ["XLSX"]:
        raise Exception("Al parecer el archivo '" + file + "' no es un archivo de Excel(XLSX) válido")
    try:
        wb = load_workbook(file, read_only=True)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        data = {}
        cols = []
        columns = list(table["map"].keys())
        for r in tqdm(ws.iter_rows(), "Leyendo Excel.."):
            if r[0].row == 1:
                for colIndex in range(len(r)):
                    if r[colIndex].value in columns:
                        cols.append(colIndex)
                        data[colIndex] = []
            else:
                for colIndex in cols:
                    data[colIndex].append(r[colIndex].value)
        return data
    except Exception as e:
        raise Exception("Error al leer el archivo: "+ str(e))

def get_table(table_name):
    if not exists('tables\\' + table_name + '.json'):
        raise Exception("No existe configuración para la tabla " + table_name)
    f = open('tables\\' +table_name+ '.json', encoding='utf-8')
    data = json.load(f)
    return data

def get_catalogues(data, conn):
    catalogues = {}
    with conn.cursor() as cursor:
        for cat in data["catalogues"]:
            fields = ", ".join(cat["select"])
            sql = "SELECT " + fields + " FROM " + cat["table"]
            cursor.execute(sql)
            records = list(map(lambda x: list(x), cursor.fetchall()))
            if "transform" in cat:
                transform_records(cat, records, None)
            keyColumnIndex = cat["select"].index(cat["key"])
            catData = {}
            for row in records:
                catData[row[keyColumnIndex]] = {}
                for colIndex in range(len(cat["select"])):
                    catData[row[keyColumnIndex]][cat["select"][colIndex]] = row[colIndex]
            catalogues[cat["name"]] = catData
    return catalogues

def transform_records(data, records, catalogues=None):
    for tfield in data["transform"]:
        colIndex = data["select"].index(tfield["field"])
        action = tfield["method"]
        trans = getattr(transform, action)
        for row in records:
            row[colIndex] = trans(row[colIndex], catalogues)

def table_exists(data, cursor):
    table_name = data['table']
    sql = "SELECT column_name FROM information_schema.columns WHERE table_name='" + table_name +"'"
    cursor.execute(sql)
    records = cursor.fetchall()
    flag = True if len(records) > 0 else False
    result = flag, records
    return result

def create_table(data, cursor):
    table_name = data['table']
    sql = "CREATE TABLE " + table_name + " ($fields$)"
    fields = []
    for field in data['fields']:
       fields.append(field +" "+ (data['fields'][field]['type']+"("+str(data['fields'][field]['length'])+")" if data['fields'][field]['length'] is not None else data['fields'][field]['type']) + ( " NULL " if data['fields'][field]['nullabe'] else " NOT NULL" ))
    field = ", ".join(fields)
    sql = sql.replace('$fields$', field)
    cursor.execute(sql)
    # Se obtienen las columnas
    sql = "SELECT column_name FROM information_schema.columns WHERE table_name='" + table_name +"'"
    cursor.execute(sql)
    records = cursor.fetchall()
    return records

def drop_table(data, cursor):
    table_name = data['table']
    sql = "DROP TABLE " + table_name 
    cursor.execute(sql)

def truncate_table(data, cursor):
    table_name = data['table']
    sql = "TRUNCATE TABLE " + table_name 
    cursor.execute(sql)

def validate_table(data, records, cursor):
    cols = list(map(lambda x: x[0], records))
    missing =[]
    for field in data['fields']:
        if field not in cols:
            missing.append(field)
    if len(missing) > 0:
        drop_table(data, cursor)
        create_table(data, cursor)

def transform_row(row, rules):
    for rule in rules:
        row[rule['colIndex']] = rule['action'](row[rule['colIndex']], rule['data'])
    return row
    
def get_transform_rules(table, data):
    file_columns = list(table["map"].values())
    rules = []
    for tField in table["transform"]:
        trans = getattr(transform, tField["method"])
        rules.append( { "colIndex" : file_columns.index(tField['field']), "action": trans, "data": data }) 
    return rules

def insert(table, data, batchSize, cursor, rules=None):
    values = []
    db_columns = list(table['fields'].keys())
    excel_columns = list(data.keys())
    insert = "INSERT INTO "+table['table']+" "+"(" + ", ".join(db_columns)+") VALUES "
    row_count = range(0, len(data[excel_columns[0]]))
    for index in tqdm(row_count, desc="Creando valores del insert"):
        row = list(map(lambda x: '' if data[x][index] is None else data[x][index], excel_columns))
        if rules is not None:
           row = transform_row(row, rules)
        rowValue = "('" + "', '".join(row) + "')"
        rowValue = rowValue.replace("'NULL'", "NULL")
        values.append(rowValue)
        if len(values) == batchSize:
            sql = insert + ", ".join(values)
            cursor.execute(sql)
            values = []
    if len(values) > 0:
        sql = insert + ", ".join(values)
        cursor.execute(sql)
        values = []

def get_extension(file):
    split = os.path.splitext(file)
    if len(split)>1:
        extension = split[1].upper()[1:]
    else:
        raise Exception("El archivo '"+file+"' no tiene extensión")
    return extension
